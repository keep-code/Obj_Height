import re
import pandas as pd
from collections import defaultdict, OrderedDict
import numpy as np
import os
from pathlib import Path


def parse_log_line(line):
    """解析单行日志数据"""
    pattern = r'zhao_od_ID:(\d+), ObjName:([^,]+), HeightLabel:(\d+), OdDir:(\d+), P1\(([^)]+)\), P2\(([^)]+)\), HeightProb:(\d+), HeightStatus:(\d+), Dist:(\d+), PosDeDis1:(\d+), l_TrainResult:(\d+), probability:(\d+)'
    match = re.match(pattern, line.strip())

    if match:
        return {
            'zhao_od_ID': int(match.group(1)),
            'ObjName': match.group(2),
            'HeightLabel': int(match.group(3)),
            'OdDir': int(match.group(4)),
            'P1': match.group(5),
            'P2': match.group(6),
            'HeightProb': int(match.group(7)),
            'HeightStatus': int(match.group(8)),
            'Dist': int(match.group(9)),
            'PosDeDis1': int(match.group(10)),
            'l_TrainResult': int(match.group(11)),
            'probability': int(match.group(12))
        }
    return None


def get_corrected_height_status_for_high_obstacles(data_list):
    """
    针对高障碍物的特殊校对规则：
    按原始log顺序遍历，2米内出现置信度为80或连续小于160的数据，对应的校对检测状态全部设置为高
    限制条件：如果模型预测(l_TrainResult)为0（低），则不改变校对检测状态
    """
    # 创建一个字典来存储每个数据点的校对状态
    corrected_status_map = {}

    # 按原始顺序遍历
    for i, data in enumerate(data_list):
        # 首先应用基础规则
        if data['PosDeDis1'] <= 500:
            corrected_status_map[i] = 2  # 设为高
        else:
            corrected_status_map[i] = data['HeightStatus']

    # 针对高障碍物的特殊规则
    i = 0
    while i < len(data_list):
        data = data_list[i]

        # 检查是否在2米（2000mm）以内
        if data['Dist'] <= 5000:
            # 检查是否出现置信度为80
            if data['HeightProb'] == 80:
                # 新增限制：如果模型预测为0（低），则不改变状态
                if data['l_TrainResult'] != 0:
                    corrected_status_map[i] = 2  # 设为高
                i += 1
                continue

            # 检查是否开始连续小于160的序列
            if data['HeightProb'] < 160:
                # 找到连续小于160的所有数据点
                start_idx = i
                while i < len(data_list) and data_list[i]['Dist'] <= 5000 and data_list[i]['HeightProb'] < 160:
                    # 新增限制：如果模型预测为0（低），则不改变状态
                    if data_list[i]['l_TrainResult'] != 0:
                        corrected_status_map[i] = 2  # 设为高
                    i += 1

                # 如果找到了连续序列，继续
                if i > start_idx:
                    continue

        i += 1

    return corrected_status_map


def get_corrected_height_status(height_status, pos_de_dis1):
    """获取校对后的检测状态（基础规则）
    当雷达距离<=500时，校对检测状态设为2(高)，否则保持原值
    """
    if pos_de_dis1 <= 500:
        return 2  # 设为高
    else:
        return height_status


def filter_valid_data(data_list):
    """过滤掉雷达距离为0的数据"""
    filtered_data = [data for data in data_list if data['PosDeDis1'] != 0]
    filtered_count = len(data_list) - len(filtered_data)
    if filtered_count > 0:
        print(f"    过滤掉雷达距离为0的数据: {filtered_count} 条")
    return filtered_data


def identify_sessions_by_distance_only(all_data, height_label):
    """
    完全重写的采样识别算法：
    基于距离变化的全局分析，识别完整的接近和远离过程
    注意：这里使用校对后的HeightStatus进行算法处理
    """
    # 先过滤掉雷达距离为0的数据
    all_data = filter_valid_data(all_data)

    if len(all_data) <= 1:
        return [all_data], [(data, 1, 'approach') for data in all_data]

    # 为每个数据点添加校对后的检测状态
    if height_label == 1:  # 高障碍物
        # 使用高障碍物的特殊校对规则
        corrected_status_map = get_corrected_height_status_for_high_obstacles(all_data)
        for idx, data in enumerate(all_data):
            data['CorrectedHeightStatus'] = corrected_status_map[idx]
    else:  # 低障碍物
        # 使用基础校对规则
        for data in all_data:
            data['CorrectedHeightStatus'] = get_corrected_height_status(data['HeightStatus'], data['PosDeDis1'])

    distances = [d['Dist'] for d in all_data]
    labeled_data = []
    approach_sessions = []

    print(f"    调试：开始分析 {len(distances)} 个数据点，距离范围 {min(distances)}-{max(distances)}")

    # 第一步：使用滑动窗口计算每个点的局部趋势
    def calculate_local_trends(distances, window_size=8):
        """计算每个点的局部趋势值"""
        trends = []

        for i in range(len(distances)):
            # 向前和向后各取window_size//2个点
            half_window = window_size // 2
            start_idx = max(0, i - half_window)
            end_idx = min(len(distances), i + half_window + 1)

            local_distances = distances[start_idx:end_idx]
            if len(local_distances) < 3:
                trends.append(0)  # 数据点太少，趋势为0
                continue

            # 计算线性回归斜率作为趋势值
            x = list(range(len(local_distances)))
            y = local_distances
            n = len(x)

            sum_x = sum(x)
            sum_y = sum(y)
            sum_xy = sum(x[i] * y[i] for i in range(n))
            sum_x2 = sum(x[i] * x[i] for i in range(n))

            # 斜率 = (n*Σxy - Σx*Σy) / (n*Σx² - (Σx)²)
            denominator = n * sum_x2 - sum_x * sum_x
            if denominator == 0:
                slope = 0
            else:
                slope = (n * sum_xy - sum_x * sum_y) / denominator

            trends.append(slope)

        return trends

    # 计算局部趋势
    trends = calculate_local_trends(distances)

    # 第二步：基于趋势识别接近和远离的阶段
    def identify_phases(trends, distances, smoothing_window=5):
        """识别接近和远离阶段"""
        # 对趋势进行平滑处理
        smoothed_trends = []
        for i in range(len(trends)):
            start = max(0, i - smoothing_window // 2)
            end = min(len(trends), i + smoothing_window // 2 + 1)
            smoothed_trends.append(sum(trends[start:end]) / (end - start))

        phases = []
        current_phase = 'approach'  # 假设开始是接近阶段
        phase_start = 0

        # 设置阈值
        TREND_THRESHOLD = 5.0  # 趋势阈值，正值表示上升，负值表示下降
        MIN_PHASE_LENGTH = 8  # 最小阶段长度

        for i in range(1, len(smoothed_trends)):
            should_switch = False

            if current_phase == 'approach':
                # 接近阶段：寻找明显的上升趋势（远离）
                if smoothed_trends[i] > TREND_THRESHOLD and i - phase_start >= MIN_PHASE_LENGTH:
                    # 检查后续几个点是否确实在上升
                    look_ahead = min(len(distances) - i, 5)
                    if look_ahead >= 3:
                        future_trend = sum(smoothed_trends[i:i + look_ahead]) / look_ahead
                        if future_trend > TREND_THRESHOLD * 0.5:
                            should_switch = True
            else:  # retreat
                # 远离阶段：寻找明显的下降趋势（接近）
                if smoothed_trends[i] < -TREND_THRESHOLD and i - phase_start >= MIN_PHASE_LENGTH:
                    # 检查后续几个点是否确实在下降
                    look_ahead = min(len(distances) - i, 5)
                    if look_ahead >= 3:
                        future_trend = sum(smoothed_trends[i:i + look_ahead]) / look_ahead
                        if future_trend < -TREND_THRESHOLD * 0.5:
                            should_switch = True

            if should_switch:
                # 记录当前阶段
                phases.append({
                    'type': current_phase,
                    'start': phase_start,
                    'end': i - 1,
                    'length': i - phase_start
                })

                # 切换到新阶段
                current_phase = 'retreat' if current_phase == 'approach' else 'approach'
                phase_start = i

        # 添加最后一个阶段
        phases.append({
            'type': current_phase,
            'start': phase_start,
            'end': len(distances) - 1,
            'length': len(distances) - phase_start
        })

        return phases

    # 识别阶段
    phases = identify_phases(trends, distances)

    print(f"    调试：识别到 {len(phases)} 个阶段")
    for i, phase in enumerate(phases):
        phase_type = "接近" if phase['type'] == 'approach' else "远离"
        start_dist = distances[phase['start']]
        end_dist = distances[phase['end']]
        print(f"      第{i + 1}个阶段：{phase_type}，长度={phase['length']}，距离={start_dist}->{end_dist}")

    # 第三步：生成标记数据和提取接近会话
    approach_count = 0
    retreat_count = 0

    for phase in phases:
        if phase['type'] == 'approach':
            approach_count += 1
            session_data = all_data[phase['start']:phase['end'] + 1]

            # 只有足够长的接近阶段才被记录为有效会话
            if len(session_data) >= 5:
                approach_sessions.append(session_data)

            # 标记数据
            for i in range(phase['start'], phase['end'] + 1):
                labeled_data.append((all_data[i], approach_count, 'approach'))
        else:  # retreat
            retreat_count += 1
            # 标记远离数据
            for i in range(phase['start'], phase['end'] + 1):
                labeled_data.append((all_data[i], retreat_count, 'retreat'))

    # 为每个接近会话按距离排序（从大到小）
    for session in approach_sessions:
        session.sort(key=lambda x: x['Dist'], reverse=True)

    print(f"    调试：提取到 {len(approach_sessions)} 个有效接近会话")

    return approach_sessions, labeled_data


def analyze_height_status_changes(session_data):
    """分析单次测试中校对后HeightStatus的变化"""
    session_data.sort(key=lambda x: x['Dist'], reverse=True)

    changes = []
    if not session_data:
        return changes

    # 使用校对后的检测状态
    current_status = session_data[0]['CorrectedHeightStatus']
    change_start_idx = 0

    for i, data in enumerate(session_data):
        if data['CorrectedHeightStatus'] != current_status:
            changes.append({
                'start_idx': change_start_idx,
                'end_idx': i - 1,
                'status': current_status,
                'start_dist': session_data[change_start_idx]['Dist'],
                'end_dist': session_data[i - 1]['Dist']
            })
            current_status = data['CorrectedHeightStatus']
            change_start_idx = i

    if change_start_idx < len(session_data):
        changes.append({
            'start_idx': change_start_idx,
            'end_idx': len(session_data) - 1,
            'status': current_status,
            'start_dist': session_data[change_start_idx]['Dist'],
            'end_dist': session_data[-1]['Dist']
        })

    return changes


def get_latest_error_distance(changes, height_label):
    """获取最近错误距离（最后一次出现错误的距离）"""
    latest_error_dist = -1  # 如果没有错误，返回-1

    # 从前往后遍历所有变化段，记录最后一个错误段的结束距离
    for change in changes:
        # 判断当前段是否为错误（基于校对后的状态）
        is_error = not ((height_label == 0 and change['status'] == 1) or
                        (height_label == 1 and change['status'] == 2))
        if is_error:
            latest_error_dist = change['end_dist']  # 修改：使用结束距离而不是起始距离

    return latest_error_dist


def get_first_error_distance(changes, height_label):
    """获取首次错误距离（第一次出现错误的距离）"""
    for change in changes:
        # 判断当前段是否为错误（基于校对后的状态）
        is_error = not ((height_label == 0 and change['status'] == 1) or
                        (height_label == 1 and change['status'] == 2))
        if is_error:
            return change['start_dist']  # 返回第一个错误段的起始距离
    return -1  # 没有错误时返回-1


def format_distance_ranges(changes, height_label):
    """格式化距离区间信息"""
    if not changes:
        return "无数据"

    wrong_ranges = []
    correct_ranges = []
    first_wrong_dist = None

    for change in changes:
        status_desc = "报高" if change['status'] == 2 else "报低"
        range_desc = f"Dist:{change['start_dist']}-{change['end_dist']}"

        # 判断是否与真实标签一致（基于校对后的状态）
        is_correct = (height_label == 0 and change['status'] == 1) or (height_label == 1 and change['status'] == 2)

        if is_correct:
            correct_ranges.append(f"{range_desc}({status_desc})")
        else:
            wrong_ranges.append(f"{range_desc}({status_desc})")
            if first_wrong_dist is None:
                first_wrong_dist = change['start_dist']

    descriptions = []
    if first_wrong_dist is not None:
        descriptions.append(f"首次错误报告距离:{first_wrong_dist}")

    if wrong_ranges:
        descriptions.append(f"错误区间:{'; '.join(wrong_ranges)}")

    if correct_ranges:
        descriptions.append(f"正确区间:{'; '.join(correct_ranges)}")

    # 简化逻辑：寻找最后一次报错后是否全程正确
    if len(changes) >= 2:  # 至少有2段才可能有转折
        # 从后往前找最后一个错误段
        last_error_end_dist = None
        for i in range(len(changes) - 1, -1, -1):
            change = changes[i]
            is_correct = (height_label == 0 and change['status'] == 1) or (height_label == 1 and change['status'] == 2)
            if not is_correct:  # 这是一个错误段
                last_error_end_dist = change['end_dist']
                break

        # 如果找到了最后一个错误段，检查之后是否全程正确
        if last_error_end_dist is not None:
            # 检查最后一个错误段之后的所有段是否都正确
            all_correct_after_error = True
            for change in changes:
                # 只检查在最后错误段之后的段
                if change['start_dist'] <= last_error_end_dist:
                    is_correct = (height_label == 0 and change['status'] == 1) or (
                            height_label == 1 and change['status'] == 2)
                    if not is_correct:
                        all_correct_after_error = False
                        break

            # 如果最后错误段之后全程正确，就添加这个描述
            if all_correct_after_error:
                expected_status = "报低" if height_label == 0 else "报高"
                descriptions.append(f"Dist<{last_error_end_dist}后全程{expected_status}")

    return "; ".join(descriptions) if descriptions else "全程正确"


def get_status_analysis(height_label, has_error, changes):
    """获取状态分析描述"""
    if not has_error:
        return "全程正确"

    for change in changes:
        # 基于校对后的状态判断错误类型
        is_error = not ((height_label == 0 and change['status'] == 1) or
                        (height_label == 1 and change['status'] == 2))
        if is_error:
            if height_label == 0 and change['status'] == 2:
                return "存在报高"
            elif height_label == 1 and change['status'] == 1:
                return "存在报低"

    return "存在错误"


def create_raw_data_sheet(labeled_data, obstacle_groups):
    """创建原始数据表，显示每条数据的分组情况，增加校对检测状态列"""
    raw_results = []

    for data, count, phase in labeled_data:
        # 跳过雷达距离为0的数据
        if data['PosDeDis1'] == 0:
            continue

        od_dir_str = "front" if data['OdDir'] == 1 else "rear"
        height_label_str = "低障碍物" if data['HeightLabel'] == 0 else "高障碍物"
        height_status_str = "高" if data['HeightStatus'] == 2 else "低"  # 原始检测状态

        # 获取校对后的检测状态（已经在数据中存储）
        corrected_status = data.get('CorrectedHeightStatus', data['HeightStatus'])
        corrected_status_str = "高" if corrected_status == 2 else "低"  # 校对检测状态

        phase_str = f"第{count}次接近" if phase == 'approach' else f"第{count}次远离"

        # 生成pack包名称
        pack_name = f"HY11-{data['ObjName']}-{data['HeightLabel']}-{od_dir_str}"

        # 基于校对后的状态判断是否正确
        is_correct = ((data['HeightLabel'] == 0 and corrected_status == 1) or
                      (data['HeightLabel'] == 1 and corrected_status == 2))

        raw_results.append({
            'pack包名称': pack_name,  # 新增列
            '障碍物名称': data['ObjName'],
            '位置': od_dir_str,
            '真实标签': height_label_str,
            '障碍物ID': data['zhao_od_ID'],
            'P1坐标': data['P1'],
            'P2坐标': data['P2'],
            '距离Dist': data['Dist'],
            '雷达距离': data['PosDeDis1'],
            '检测状态': height_status_str,  # 原始检测状态
            '校对检测状态': corrected_status_str,  # 新增：校对后的检测状态
            '置信度': data['HeightProb'],
            '模型预测': data['l_TrainResult'],
            '模型概率': data['probability'],
            '检测阶段': phase_str,
            '是否正确': "✓" if is_correct else "✗"  # 基于校对后状态判断
        })

    return pd.DataFrame(raw_results)


def ensure_directory_exists(file_path):
    """确保文件所在目录存在"""
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        try:
            os.makedirs(directory, exist_ok=True)
            print(f"✅ 创建目录: {directory}")
        except Exception as e:
            print(f"❌ 创建目录失败: {e}")
            return False
    return True


def save_to_excel(detailed_df, summary_df, raw_df, base_path, include_raw=False):
    """保存为Excel格式，自动调整列宽"""
    excel_path = f"{base_path}_分析结果.xlsx"

    if not ensure_directory_exists(excel_path):
        return None

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            detailed_df.to_excel(writer, sheet_name='详细分析', index=False)
            summary_df.to_excel(writer, sheet_name='简洁汇总', index=False)

            if include_raw and raw_df is not None:
                raw_df.to_excel(writer, sheet_name='原始数据分组', index=False)

            workbook = writer.book
            detailed_sheet = workbook['详细分析']
            adjust_column_width(detailed_sheet, detailed_df)

            summary_sheet = workbook['简洁汇总']
            adjust_column_width(summary_sheet, summary_df)

            if include_raw and raw_df is not None:
                raw_sheet = workbook['原始数据分组']
                adjust_column_width(raw_sheet, raw_df)

            format_excel_sheets(workbook)

        sheet_info = "包含原始数据分组表" if include_raw else "不包含原始数据表"
        print(f"✅ Excel文件已保存: {excel_path} ({sheet_info})")
        return excel_path

    except ImportError:
        print("⚠️  警告: 未安装 openpyxl 库，无法生成Excel文件")
        return None
    except Exception as e:
        print(f"❌ 保存Excel文件时出错: {e}")
        return None


def adjust_column_width(sheet, df):
    """自动调整列宽"""
    try:
        from openpyxl.utils import get_column_letter

        for col_idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(column))

            for value in df[column]:
                if pd.notna(value):
                    str_value = str(value)
                    length = len(str_value) + sum(1 for char in str_value if ord(char) > 127)
                    max_length = max(max_length, length)

            adjusted_width = min(max(max_length + 2, 10), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    except ImportError:
        print("⚠️  警告: 无法调整列宽，请确保安装了 openpyxl 库")
    except Exception as e:
        print(f"⚠️  调整列宽时出现问题: {e}")


def format_excel_sheets(workbook):
    """格式化Excel工作表"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            sheet.freeze_panes = 'A2'

    except ImportError:
        print("⚠️  警告: 无法应用Excel格式化")
    except Exception as e:
        print(f"⚠️  格式化Excel时出现问题: {e}")


def save_to_csv(detailed_df, summary_df, raw_df, base_path, include_raw=False):
    """保存为CSV格式"""
    detailed_csv = f"{base_path}_详细分析.csv"
    summary_csv = f"{base_path}_简洁汇总.csv"
    raw_csv = f"{base_path}_原始数据分组.csv"

    if not ensure_directory_exists(detailed_csv) or not ensure_directory_exists(summary_csv):
        return None, None, None

    try:
        detailed_df.to_csv(detailed_csv, index=False, encoding='utf-8-sig')
        summary_df.to_csv(summary_csv, index=False, encoding='utf-8-sig')

        saved_files = [detailed_csv, summary_csv]

        if include_raw and raw_df is not None:
            raw_df.to_csv(raw_csv, index=False, encoding='utf-8-sig')
            saved_files.append(raw_csv)

        print(f"✅ CSV文件已保存:")
        print(f"  📁 详细分析: {detailed_csv}")
        print(f"  📊 简洁汇总: {summary_csv}")
        if include_raw and raw_df is not None:
            print(f"  📋 原始数据分组: {raw_csv}")

        return tuple(saved_files)

    except Exception as e:
        print(f"❌ 保存CSV文件时出错: {e}")
        return None, None, None


def analyze_obstacle_data(log_file_path, output_path, output_format='excel', include_raw_data=False):
    """
    主要分析函数
    """
    if not os.path.exists(log_file_path):
        print(f"❌ 错误: 日志文件不存在: {log_file_path}")
        return None, None, None

    print(f"📂 开始分析日志文件: {log_file_path}")
    print(f"🔧 原始数据分组表: {'启用' if include_raw_data else '禁用'}")
    print(f"🔧 校对检测状态规则: ")
    print(f"   - 基础规则：雷达距离<=500时，检测状态设为'高'")
    print(f"   - 高障碍物特殊规则：2米内置信度为80或连续<160时，检测状态设为'高'")
    print(f"🔧 数据过滤规则: 忽略雷达距离为0的数据")

    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        print(f"📝 读取到 {len(lines)} 行数据")
    except Exception as e:
        print(f"❌ 读取日志文件失败: {e}")
        return None, None, None

    all_data = []
    failed_lines = 0

    for line_num, line in enumerate(lines, 1):
        parsed = parse_log_line(line)
        if parsed:
            all_data.append(parsed)
        else:
            failed_lines += 1
            if failed_lines <= 3:
                print(f"⚠️  第{line_num}行解析失败: {line.strip()[:100]}...")

    print(f"✅ 成功解析 {len(all_data)} 条数据，失败 {failed_lines} 条")

    if not all_data:
        print("❌ 错误: 没有找到有效的日志数据")
        return None, None, None

    obstacle_groups = defaultdict(list)
    for data in all_data:
        key = (data['ObjName'], data['OdDir'], data['HeightLabel'])
        obstacle_groups[key].append(data)

    print(f"🔍 识别到 {len(obstacle_groups)} 个不同的障碍物组合")

    detailed_results = []
    summary_results = []
    all_labeled_data = []

    for (obj_name, od_dir, height_label), data_list in obstacle_groups.items():
        od_dir_str = "front" if od_dir == 1 else "rear"
        height_label_str = "低障碍物" if height_label == 0 else "高障碍物"

        print(f"\n🔍 分析障碍物: {obj_name} ({od_dir_str}, {height_label_str})")

        approach_sessions, labeled_data = identify_sessions_by_distance_only(data_list, height_label)
        all_labeled_data.extend(labeled_data)

        print(f"  识别出 {len(approach_sessions)} 次接近过程")

        id_groups = defaultdict(list)
        for session_idx, session in enumerate(approach_sessions, 1):
            for data in session:
                id_groups[data['zhao_od_ID']].append((data, session_idx))

        print(f"  涉及 {len(id_groups)} 个不同的障碍物ID: {list(id_groups.keys())}")

        obstacle_has_any_error = False
        total_sessions_all_ids = len(approach_sessions)
        total_points_all_ids = sum(len(session) for session in approach_sessions)

        # 用于记录所有最近错误距离，以计算最小错误距离
        all_latest_error_distances = []
        # 用于记录所有首次错误距离，以计算最大错误距离
        all_first_error_distances = []

        for zhao_od_id, id_data_with_session in id_groups.items():
            print(f"  分析ID {zhao_od_id}: {len(id_data_with_session)} 个数据点")

            session_groups = defaultdict(list)
            for data, session_idx in id_data_with_session:
                session_groups[session_idx].append(data)

            all_sessions_correct = True
            incorrect_sessions = []

            for session_idx, session_data in session_groups.items():
                # 基于校对后的状态判断是否正确
                session_correct = all(
                    (height_label == 0 and data.get('CorrectedHeightStatus',
                                                    get_corrected_height_status(data['HeightStatus'],
                                                                                data['PosDeDis1'])) == 1) or
                    (height_label == 1 and data.get('CorrectedHeightStatus',
                                                    get_corrected_height_status(data['HeightStatus'],
                                                                                data['PosDeDis1'])) == 2)
                    for data in session_data
                )

                if not session_correct:
                    all_sessions_correct = False
                    incorrect_sessions.append(session_idx)
                    obstacle_has_any_error = True

            if all_sessions_correct:
                total_points = sum(len(session_data) for session_data in session_groups.values())
                all_distances = [data['Dist'] for session_data in session_groups.values() for data in session_data]
                min_dist = min(all_distances)
                max_dist = max(all_distances)

                detailed_results.append({
                    '障碍物名称': obj_name,
                    '位置': od_dir_str,
                    '真实标签': height_label_str,
                    '障碍物ID': zhao_od_id,
                    '总测试次数': len(session_groups),
                    '测试次数说明': f"第{','.join(map(str, sorted(session_groups.keys())))}次" if len(
                        session_groups) > 1 else "第1次",
                    '数据点总数': total_points,
                    '距离范围(mm)': f"{max_dist}~{min_dist}",
                    '检测状态': "全程正确",
                    '首次错误距离(mm)': "-1",  # 修改：无错误时显示-1
                    '最近错误距离(mm)': "-1",  # 新增列：无错误时为-1
                    '错误详情': "无",
                    '备注': f"共{len(session_groups)}次测试，全程校对后HeightStatus与HeightLabel一致"
                })

            else:
                # 对于错误的次数，分别输出
                for session_idx in incorrect_sessions:
                    session_data = session_groups[session_idx]
                    # 为session_data中每个数据添加校对后的状态（如果还没有的话）
                    if 'CorrectedHeightStatus' not in session_data[0]:
                        if height_label == 1:  # 高障碍物
                            corrected_status_map = get_corrected_height_status_for_high_obstacles(session_data)
                            for idx, data in enumerate(session_data):
                                data['CorrectedHeightStatus'] = corrected_status_map[idx]
                        else:  # 低障碍物
                            for data in session_data:
                                data['CorrectedHeightStatus'] = get_corrected_height_status(data['HeightStatus'],
                                                                                            data['PosDeDis1'])

                    changes = analyze_height_status_changes(session_data)
                    error_desc = format_distance_ranges(changes, height_label)
                    status_analysis = get_status_analysis(height_label, True, changes)

                    # 获取首次错误距离
                    first_error_dist = get_first_error_distance(changes, height_label)

                    # 获取最近错误距离
                    latest_error_dist = get_latest_error_distance(changes, height_label)

                    # 记录错误距离用于汇总计算
                    if first_error_dist > 0:
                        all_first_error_distances.append(first_error_dist)
                    if latest_error_dist > 0:
                        all_latest_error_distances.append(latest_error_dist)

                    detailed_results.append({
                        '障碍物名称': obj_name,
                        '位置': od_dir_str,
                        '真实标签': height_label_str,
                        '障碍物ID': zhao_od_id,
                        '总测试次数': len(session_groups),
                        '测试次数说明': f"第{session_idx}次",
                        '数据点总数': len(session_data),
                        '距离范围(mm)': f"{max(d['Dist'] for d in session_data)}~{min(d['Dist'] for d in session_data)}",
                        '检测状态': status_analysis,
                        '首次错误距离(mm)': str(first_error_dist) if first_error_dist > 0 else "-1",  # 修改：无错误时显示-1
                        '最近错误距离(mm)': str(latest_error_dist) if latest_error_dist > 0 else "-1",  # 新增列
                        '错误详情': error_desc,
                        '备注': f"第{session_idx}次测试中校对后HeightStatus与HeightLabel不一致"
                    })

                # 对于正确的次数，合并输出一条记录
                correct_sessions = [i for i in session_groups.keys() if i not in incorrect_sessions]
                if correct_sessions:
                    correct_data = [session_groups[i] for i in correct_sessions]
                    total_points = sum(len(session_data) for session_data in correct_data)
                    all_distances = [data['Dist'] for session_data in correct_data for data in session_data]
                    min_dist = min(all_distances)
                    max_dist = max(all_distances)

                    detailed_results.append({
                        '障碍物名称': obj_name,
                        '位置': od_dir_str,
                        '真实标签': height_label_str,
                        '障碍物ID': zhao_od_id,
                        '总测试次数': len(session_groups),
                        '测试次数说明': f"第{','.join(map(str, correct_sessions))}次",
                        '数据点总数': total_points,
                        '距离范围(mm)': f"{max_dist}~{min_dist}",
                        '检测状态': "全程正确",
                        '首次错误距离(mm)': "-1",  # 修改：无错误时显示-1
                        '最近错误距离(mm)': "-1",  # 新增列：无错误时为-1
                        '错误详情': "无",
                        '备注': f"第{','.join(map(str, correct_sessions))}次测试全程正确"
                    })

        # 生成简洁版结果
        if not obstacle_has_any_error:
            all_distances = [data['Dist'] for data in data_list]
            # 生成pack包名称
            od_dir_pack_str = "front" if od_dir == 1 else "rear"
            pack_name = f"HY11-{obj_name}-{height_label}-{od_dir_pack_str}"

            # 根据真实标签生成备注
            if height_label == 0:  # 低障碍物
                remark = f"重复{total_sessions_all_ids}次，未报高"
            else:  # 高障碍物
                remark = f"重复{total_sessions_all_ids}次，未报低"

            summary_results.append({
                'pack包名称': pack_name,  # 新增列
                '障碍物名称': obj_name,
                '位置': od_dir_str,
                '真实标签': height_label_str,
                '总测试次数': total_sessions_all_ids,
                '涉及ID数量': len(id_groups),
                '数据点总数': total_points_all_ids,
                '距离范围(mm)': f"{max(all_distances)}~{min(all_distances)}",
                '检测结果': "✓ 全程正确",
                '最大错误距离(mm)': "-1",  # 新增列：无错误时为-1
                '最小错误距离(mm)': "-1",  # 新增列：无错误时为-1
                '备注': remark
            })
        else:
            error_type = "存在报高" if height_label == 0 else "存在报低"
            all_distances = [data['Dist'] for data in data_list]
            # 生成pack包名称
            od_dir_pack_str = "front" if od_dir == 1 else "rear"
            pack_name = f"HY11-{obj_name}-{height_label}-{od_dir_pack_str}"

            # 计算最大和最小错误距离
            max_error_dist = max(all_first_error_distances) if all_first_error_distances else -1
            min_error_dist = min(all_latest_error_distances) if all_latest_error_distances else -1

            # 根据真实标签生成备注
            if height_label == 0:  # 低障碍物
                if max_error_dist > 0 and min_error_dist > 0:
                    remark = f"重复{total_sessions_all_ids}次，存在报高，最远{max_error_dist}，最近{min_error_dist}"
                else:
                    remark = f"重复{total_sessions_all_ids}次，存在报高"
            else:  # 高障碍物
                if max_error_dist > 0 and min_error_dist > 0:
                    remark = f"重复{total_sessions_all_ids}次，存在报低，最远{max_error_dist}，最近{min_error_dist}"
                else:
                    remark = f"重复{total_sessions_all_ids}次，存在报低"

            summary_results.append({
                'pack包名称': pack_name,  # 新增列
                '障碍物名称': obj_name,
                '位置': od_dir_str,
                '真实标签': height_label_str,
                '总测试次数': total_sessions_all_ids,
                '涉及ID数量': len(id_groups),
                '数据点总数': total_points_all_ids,
                '距离范围(mm)': f"{max(all_distances)}~{min(all_distances)}",
                '检测结果': f"✗ {error_type}",
                '最大错误距离(mm)': str(max_error_dist) if max_error_dist > 0 else "-1",  # 新增列
                '最小错误距离(mm)': str(min_error_dist) if min_error_dist > 0 else "-1",  # 新增列
                '备注': remark
            })

    if not detailed_results:
        print("❌ 错误: 没有生成有效的分析结果")
        return None, None, None

    detailed_df = pd.DataFrame(detailed_results)
    summary_df = pd.DataFrame(summary_results)

    raw_df = None
    if include_raw_data:
        raw_df = create_raw_data_sheet(all_labeled_data, obstacle_groups)
        print(f"📋 生成原始数据分组表: {len(raw_df)}条记录")

    print(f"\n📊 生成分析结果: 详细记录{len(detailed_results)}条，汇总记录{len(summary_results)}条")

    if output_path.endswith('.csv') or output_path.endswith('.xlsx'):
        base_path = os.path.splitext(output_path)[0]
    else:
        base_path = output_path

    saved_files = []

    if output_format.lower() in ['excel', 'both']:
        excel_file = save_to_excel(detailed_df, summary_df, raw_df, base_path, include_raw_data)
        if excel_file:
            saved_files.append(excel_file)

    if output_format.lower() in ['csv', 'both']:
        csv_files = save_to_csv(detailed_df, summary_df, raw_df, base_path, include_raw_data)
        if csv_files and csv_files[0]:
            saved_files.extend([f for f in csv_files if f])

    print("=" * 80)
    print("🎯 障碍物检测日志分析完成!")
    print("=" * 80)

    if saved_files:
        print("📁 输出文件:")
        for file in saved_files:
            print(f"  • {file}")
    else:
        print("❌ 警告: 没有成功保存任何文件")

    print()
    print("📈 统计概览:")
    print("-" * 50)
    print(f"🔍 分析障碍物类型总数: {len(obstacle_groups)}")
    print(f"📝 详细记录条数: {len(detailed_results)}")
    print(f"📋 汇总记录条数: {len(summary_results)}")
    print(f"✅ 全程正确的障碍物: {len(summary_df[summary_df['检测结果'].str.contains('全程正确')])}")
    print(f"❌ 存在错误的障碍物: {len(summary_df[summary_df['检测结果'].str.contains('存在')])}")
    if include_raw_data and raw_df is not None:
        print(f"📋 原始数据记录条数: {len(raw_df)}")

    print("\n📋 简洁汇总表预览:")
    print("-" * 50)
    if len(summary_df) > 0:
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 30)
        print(summary_df.head(10).to_string(index=False))
        if len(summary_df) > 10:
            print(f"\n... 还有 {len(summary_df) - 10} 行数据，详见输出文件")
    else:
        print("无数据")

    print("\n" + "=" * 80)

    print("🆕 数据过滤说明:")
    print("-" * 50)
    print("🔧 核心改进:")
    print("  • 自动过滤雷达距离为0的数据，这些数据不参与错误识别")
    print("  • 在原始数据分组表中也会自动排除雷达距离为0的记录")
    print("  • 确保分析结果的准确性和可靠性")

    print("\n🆕 高障碍物特殊校对规则:")
    print("-" * 50)
    print("🔧 置信度校对增强:")
    print("  • 对于真实标签为高的障碍物，增加置信度校对规则")
    print("  • 在2米（2000mm）以内:")
    print("    - 出现置信度为80的数据，校对检测状态设为'高'")
    print("    - 连续出现置信度<160的数据，校对检测状态全部设为'高'")
    print("  • 此规则仅适用于高障碍物，低障碍物保持原有校对规则")

    print("\n🆕 显示格式优化:")
    print("-" * 50)
    print("🔧 详细分析表改进:")
    print("  • '首次错误距离(mm)'：无错误时显示'-1'而非'无'")
    print("  • 统一数值格式，便于数据处理和排序")

    print("\n🆕 简洁汇总表新增列:")
    print("-" * 50)
    print("🔧 距离统计增强:")
    print("  • '最大错误距离(mm)'：该障碍物所有测试中首次错误距离的最大值")
    print("    - 计算来源：详细分析表中所有'首次错误距离(mm)' > 0的最大值")
    print("    - 反映最远的错误发生位置")
    print("  • '最小错误距离(mm)'：该障碍物所有测试中最近错误距离的最小值")
    print("    - 计算来源：详细分析表中所有'最近错误距离(mm)' > 0的最小值")
    print("    - 反映最近的错误发生位置")

    print("\n🆕 备注格式优化:")
    print("-" * 50)
    print("🔧 备注内容标准化:")
    print("  • 低障碍物无错误：'重复X次，未报高'")
    print("  • 低障碍物有错误：'重复X次，存在报高，最远Y，最近Z'")
    print("  • 高障碍物无错误：'重复X次，未报低'")
    print("  • 高障碍物有错误：'重复X次，存在报低，最远Y，最近Z'")
    print("  • X = 总测试次数，Y = 最大错误距离，Z = 最小错误距离")

    print("\n💡 算法保持不变:")
    print("-" * 50)
    print("🔧 核心逻辑:")
    print("  • 距离过滤：过滤雷达距离为0的数据")
    print("  • 状态校对：")
    print("    - 基础规则：雷达距离<=500时检测状态设为'高'")
    print("    - 高障碍物特殊规则：2米内特定置信度条件下设为'高'")
    print("  • 会话识别：基于线性回归的趋势分析")
    print("  • 错误统计：基于校对后状态的准确判断")

    return detailed_df, summary_df, raw_df


# 使用示例
if __name__ == "__main__":
    log_file_path = r"D:\PythonProject\data\log_files\1_40.log"
    output_path = r"D:\PythonProject\data\csv_files\6\1_40"

    try:
        print("🚀 开始校对检测状态的日志分析...")
        print(f"📂 日志文件: {log_file_path}")
        print(f"📁 输出路径: {output_path}")
        print("-" * 80)

        include_raw_data = True

        detailed_df, summary_df, raw_df = analyze_obstacle_data(
            log_file_path,
            output_path,
            output_format='excel',
            include_raw_data=include_raw_data
        )

        if detailed_df is not None and summary_df is not None:
            print("\n🔍 详细分析表字段说明:")
            print("-" * 50)
            field_descriptions = {
                '障碍物名称': '障碍物类型名称',
                '位置': 'front(车前) / rear(车后)',
                '真实标签': '实际的高低标签',
                '障碍物ID': 'zhao_od_ID编号',
                '总测试次数': '该ID在所有接近过程中的测试次数',
                '测试次数说明': '当前记录对应的测试次数',
                '数据点总数': '数据记录条数',
                '距离范围(mm)': '测试距离范围(最远~最近)',
                '检测状态': '全程正确/存在报高/存在报低(基于校对后状态)',
                '首次错误距离(mm)': '第一次出现错误时的距离(基于校对后状态，-1表示无错误)',
                '最近错误距离(mm)': '最后一次出现错误时的距离(基于校对后状态，-1表示无错误)',
                '错误详情': '错误的具体距离区间信息(基于校对后状态)',
                '备注': '额外说明信息'
            }

            for field, desc in field_descriptions.items():
                print(f"  • {field}: {desc}")

            print("\n📊 简洁汇总表字段说明:")
            print("-" * 50)
            summary_field_descriptions = {
                'pack包名称': 'HY11-障碍物名称-标签-位置格式',
                '障碍物名称': '障碍物类型名称',
                '位置': 'front(车前) / rear(车后)',
                '真实标签': '实际的高低标签',
                '总测试次数': '该障碍物的总测试次数',
                '涉及ID数量': '涉及的不同障碍物ID数量',
                '数据点总数': '总数据记录条数',
                '距离范围(mm)': '测试距离范围(最远~最近)',
                '检测结果': '✓全程正确 / ✗存在报高 / ✗存在报低',
                '最大错误距离(mm)': '所有测试中首次错误距离的最大值(-1表示无错误)',
                '最小错误距离(mm)': '所有测试中最近错误距离的最小值(-1表示无错误)',
                '备注': '标准化格式的测试结果描述'
            }

            for field, desc in summary_field_descriptions.items():
                print(f"  • {field}: {desc}")

            if include_raw_data and raw_df is not None:
                print("\n📋 原始数据分组表字段说明:")
                print("-" * 50)
                raw_field_descriptions = {
                    'pack包名称': 'HY11-障碍物名称-标签-位置格式',
                    '障碍物名称': '障碍物类型名称',
                    '位置': 'front(车前) / rear(车后)',
                    '真实标签': '实际的高低标签',
                    '障碍物ID': 'zhao_od_ID编号',
                    'P1坐标': '障碍物坐标点1',
                    'P2坐标': '障碍物坐标点2',
                    '距离Dist': '车辆到障碍物的距离',
                    '雷达距离': '雷达直接回波距离(已过滤0值)',
                    '检测状态': '系统报告的原始高低状态',
                    '校对检测状态': '基于雷达距离和置信度校对后的高低状态',
                    '置信度': '系统置信度',
                    '模型预测': '模型预测结果',
                    '模型概率': '模型预测概率值',
                    '检测阶段': '第几次接近或远离过程',
                    '是否正确': '校对后检测结果是否与真实标签一致'
                }

                for field, desc in raw_field_descriptions.items():
                    print(f"  • {field}: {desc}")

        else:
            print("\n❌ 分析失败，请检查:")
            print("  1. 日志文件是否存在且格式正确")
            print("  2. 输出目录是否有写入权限")
            print("  3. 是否安装了必要的依赖库")

    except FileNotFoundError:
        print(f"❌ 错误: 找不到日志文件 '{log_file_path}'")
        print("请确保文件路径正确，或将日志内容保存为对应的日志文件")
    except Exception as e:
        print(f"❌ 分析过程中出现错误: {e}")
        import traceback

        traceback.print_exc()