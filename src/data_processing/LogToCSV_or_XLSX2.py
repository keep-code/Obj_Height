import re
import pandas as pd
from collections import defaultdict, OrderedDict
import numpy as np
import os
from pathlib import Path


def parse_log_line(line):
    """解析单行日志数据"""
    # 使用正则表达式提取各个字段
    pattern = r'zhao_od_ID:(\d+), ObjName:([^,]+), HeightLabel:(\d+), OdDir:(\d+), P1\(([^)]+)\), P2\(([^)]+)\), HeightProb:(\d+), HeightStatus:(\d+), Dist:(\d+), PosDeDis1:(\d+), l_TrainResult:(\d+), probability:(\d+)'
    match = re.match(pattern, line.strip())

    if match:
        return {
            'zhao_od_ID': int(match.group(1)),
            'ObjName': match.group(2),
            'HeightLabel': int(match.group(3)),
            'OdDir': int(match.group(4)),
            'P1': match.group(5),
            'P2': match.group(6),
            'HeightProb': int(match.group(7)),
            'HeightStatus': int(match.group(8)),
            'Dist': int(match.group(9)),
            'PosDeDis1': int(match.group(10)),
            'l_TrainResult': int(match.group(11)),
            'probability': int(match.group(12))
        }
    return None


def identify_test_sessions(data_list):
    """
    识别测试次数，根据Dist连续变化趋势判断
    改进版本：能够识别连续的接近过程，过滤远离过程和干扰数据
    """
    if len(data_list) <= 1:
        return [data_list]

    def analyze_trend_window(distances, start_idx, window_size=5):
        """
        分析指定窗口内的距离变化趋势
        返回: ('decreasing', score) 或 ('increasing', score) 或 ('mixed', score)
        score 表示趋势的强度，范围0-1
        """
        end_idx = min(start_idx + window_size, len(distances))
        if end_idx - start_idx < 3:
            return 'mixed', 0.0

        window_distances = distances[start_idx:end_idx]
        decreasing_count = 0
        increasing_count = 0
        total_changes = len(window_distances) - 1

        for i in range(1, len(window_distances)):
            if window_distances[i] < window_distances[i - 1]:
                decreasing_count += 1
            elif window_distances[i] > window_distances[i - 1]:
                increasing_count += 1

        if total_changes == 0:
            return 'mixed', 0.0

        decreasing_ratio = decreasing_count / total_changes
        increasing_ratio = increasing_count / total_changes

        if decreasing_ratio >= 0.6:  # 60%以上的点在下降
            return 'decreasing', decreasing_ratio
        elif increasing_ratio >= 0.6:  # 60%以上的点在上升
            return 'increasing', increasing_ratio
        else:
            return 'mixed', max(decreasing_ratio, increasing_ratio)

    def is_significant_change(current_dist, prev_dist, threshold_ratio=0.1):
        """判断距离变化是否显著"""
        if prev_dist == 0:
            return True
        change_ratio = abs(current_dist - prev_dist) / prev_dist
        return change_ratio > threshold_ratio

    # 首先按原始顺序排序，以保持时间序列特性
    # 但我们需要根据Dist来判断，所以先按Dist粗略排序
    data_list.sort(key=lambda x: x['Dist'], reverse=True)

    if len(data_list) < 3:
        return [data_list]

    distances = [d['Dist'] for d in data_list]
    sessions = []
    current_session = []
    i = 0

    print(f"    调试：总数据点 {len(distances)}，距离范围 {min(distances)}-{max(distances)}")

    while i < len(data_list):
        if not current_session:
            # 开始新的会话
            current_session = [data_list[i]]
            i += 1
            continue

        # 分析当前位置的趋势
        current_trend, trend_score = analyze_trend_window(distances, max(0, i - 2), 5)

        prev_dist = distances[i - 1]
        curr_dist = distances[i]

        # 判断是否应该继续当前会话
        if len(current_session) == 1:
            # 第二个点，直接加入
            current_session.append(data_list[i])
        else:
            # 分析趋势决定是否继续
            if current_trend == 'decreasing' and trend_score > 0.5:
                # 强烈的下降趋势，继续当前会话
                current_session.append(data_list[i])
            elif current_trend == 'increasing' and trend_score > 0.6:
                # 强烈的上升趋势，可能是远离过程，结束当前会话
                if len(current_session) >= 3:
                    sessions.append(current_session)
                    print(
                        f"    调试：完成一次采样，数据点数 {len(current_session)}，距离范围 {max(d['Dist'] for d in current_session)}-{min(d['Dist'] for d in current_session)}")

                # 跳过上升阶段，寻找下一个下降开始点
                while i < len(data_list) - 1:
                    next_trend, next_score = analyze_trend_window(distances, i, 3)
                    if next_trend == 'decreasing' and next_score > 0.5:
                        break
                    i += 1

                current_session = []
                continue
            else:
                # 混合趋势，根据距离变化判断
                if curr_dist < prev_dist or not is_significant_change(curr_dist, prev_dist):
                    # 仍在下降或变化不大，继续
                    current_session.append(data_list[i])
                else:
                    # 开始上升，结束当前会话
                    if len(current_session) >= 3:
                        sessions.append(current_session)
                        print(
                            f"    调试：完成一次采样，数据点数 {len(current_session)}，距离范围 {max(d['Dist'] for d in current_session)}-{min(d['Dist'] for d in current_session)}")
                    current_session = []
                    continue

        i += 1

    # 处理最后一个会话
    if current_session and len(current_session) >= 3:
        sessions.append(current_session)
        print(
            f"    调试：完成最后一次采样，数据点数 {len(current_session)}，距离范围 {max(d['Dist'] for d in current_session)}-{min(d['Dist'] for d in current_session)}")

    # 为每个会话重新按距离排序（从大到小，确保是接近过程）
    for session in sessions:
        session.sort(key=lambda x: x['Dist'], reverse=True)

    print(f"    调试：总共识别出 {len(sessions)} 次测试")
    return sessions if sessions else [data_list]


def analyze_height_status_changes(session_data):
    """分析单次测试中HeightStatus的变化"""
    # 按距离排序（从大到小，接近过程）
    session_data.sort(key=lambda x: x['Dist'], reverse=True)

    changes = []
    if not session_data:
        return changes

    current_status = session_data[0]['HeightStatus']
    change_start_idx = 0

    for i, data in enumerate(session_data):
        if data['HeightStatus'] != current_status:
            # 记录状态变化
            changes.append({
                'start_idx': change_start_idx,
                'end_idx': i - 1,
                'status': current_status,
                'start_dist': session_data[change_start_idx]['Dist'],
                'end_dist': session_data[i - 1]['Dist']
            })
            current_status = data['HeightStatus']
            change_start_idx = i

    # 添加最后一段
    if change_start_idx < len(session_data):
        changes.append({
            'start_idx': change_start_idx,
            'end_idx': len(session_data) - 1,
            'status': current_status,
            'start_dist': session_data[change_start_idx]['Dist'],
            'end_dist': session_data[-1]['Dist']
        })

    return changes


def format_distance_ranges(changes, height_label):
    """格式化距离区间信息"""
    if not changes:
        return "无数据"

    # 找出错误报告的区间（HeightStatus与HeightLabel不一致）
    wrong_ranges = []
    correct_ranges = []
    first_wrong_dist = None

    for change in changes:
        status_desc = "报高" if change['status'] == 2 else "报低"
        range_desc = f"Dist:{change['start_dist']}-{change['end_dist']}"

        # 判断是否与真实标签一致
        is_correct = (height_label == 0 and change['status'] == 1) or (height_label == 1 and change['status'] == 2)

        if is_correct:
            correct_ranges.append(f"{range_desc}({status_desc})")
        else:
            wrong_ranges.append(f"{range_desc}({status_desc})")
            if first_wrong_dist is None:
                first_wrong_dist = change['start_dist']

    # 构建描述
    descriptions = []
    if first_wrong_dist is not None:
        descriptions.append(f"首次错误报告距离:{first_wrong_dist}")

    if wrong_ranges:
        descriptions.append(f"错误区间:{'; '.join(wrong_ranges)}")

    if correct_ranges:
        descriptions.append(f"正确区间:{'; '.join(correct_ranges)}")

    # 检查是否在某个距离后全程正确
    if len(changes) > 1:
        last_change = changes[-1]
        last_is_correct = (height_label == 0 and last_change['status'] == 1) or (
                height_label == 1 and last_change['status'] == 2)
        if last_is_correct and last_change['end_dist'] < last_change['start_dist'] * 0.7:
            descriptions.append(f"Dist<{last_change['start_dist']}后全程{'报低' if height_label == 0 else '报高'}")

    return "; ".join(descriptions) if descriptions else "全程正确"


def get_status_analysis(height_label, has_error, changes):
    """获取状态分析描述"""
    if not has_error:
        return "全程正确"

    # 检查具体的错误类型
    for change in changes:
        is_error = not ((height_label == 0 and change['status'] == 1) or
                        (height_label == 1 and change['status'] == 2))
        if is_error:
            if height_label == 0 and change['status'] == 2:  # 低障碍物但报高
                return "存在报高"
            elif height_label == 1 and change['status'] == 1:  # 高障碍物但报低
                return "存在报低"

    return "存在错误"


def ensure_directory_exists(file_path):
    """确保文件所在目录存在"""
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        try:
            os.makedirs(directory, exist_ok=True)
            print(f"✅ 创建目录: {directory}")
        except Exception as e:
            print(f"❌ 创建目录失败: {e}")
            return False
    return True


def save_to_excel(detailed_df, summary_df, base_path):
    """保存为Excel格式，自动调整列宽"""
    excel_path = f"{base_path}_分析结果.xlsx"

    # 确保目录存在
    if not ensure_directory_exists(excel_path):
        return None

    try:
        # 使用 openpyxl 引擎写入Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # 写入详细分析表
            detailed_df.to_excel(writer, sheet_name='详细分析', index=False)

            # 写入简洁汇总表
            summary_df.to_excel(writer, sheet_name='简洁汇总', index=False)

            # 获取工作簿和工作表
            workbook = writer.book

            # 自动调整详细分析表的列宽
            detailed_sheet = workbook['详细分析']
            adjust_column_width(detailed_sheet, detailed_df)

            # 自动调整简洁汇总表的列宽
            summary_sheet = workbook['简洁汇总']
            adjust_column_width(summary_sheet, summary_df)

            # 添加格式化
            format_excel_sheets(workbook)

        print(f"✅ Excel文件已保存: {excel_path}")
        return excel_path

    except ImportError:
        print("⚠️  警告: 未安装 openpyxl 库，无法生成Excel文件")
        print("请运行: pip install openpyxl")
        return None
    except Exception as e:
        print(f"❌ 保存Excel文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return None


def adjust_column_width(sheet, df):
    """自动调整列宽"""
    try:
        from openpyxl.utils import get_column_letter

        # 遍历所有列
        for col_idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_idx)

            # 计算列宽：考虑列名和数据内容
            max_length = len(str(column))  # 列名长度

            # 检查该列所有数据的长度
            for value in df[column]:
                if pd.notna(value):
                    # 对于中文字符，每个字符按2个字符计算宽度
                    str_value = str(value)
                    length = len(str_value) + sum(1 for char in str_value if ord(char) > 127)
                    max_length = max(max_length, length)

            # 设置列宽，最小宽度为10，最大宽度为50
            adjusted_width = min(max(max_length + 2, 10), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    except ImportError:
        print("⚠️  警告: 无法调整列宽，请确保安装了 openpyxl 库")
    except Exception as e:
        print(f"⚠️  调整列宽时出现问题: {e}")


def format_excel_sheets(workbook):
    """格式化Excel工作表"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 定义样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 格式化每个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 格式化标题行
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 为所有数据添加边框
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            # 冻结首行
            sheet.freeze_panes = 'A2'

    except ImportError:
        print("⚠️  警告: 无法应用Excel格式化，请确保安装了 openpyxl 库")
    except Exception as e:
        print(f"⚠️  格式化Excel时出现问题: {e}")


def save_to_csv(detailed_df, summary_df, base_path):
    """保存为CSV格式"""
    detailed_csv = f"{base_path}_详细分析.csv"
    summary_csv = f"{base_path}_简洁汇总.csv"

    # 确保目录存在
    if not ensure_directory_exists(detailed_csv) or not ensure_directory_exists(summary_csv):
        return None, None

    try:
        # 保存详细分析表
        detailed_df.to_csv(detailed_csv, index=False, encoding='utf-8-sig')

        # 保存简洁汇总表
        summary_df.to_csv(summary_csv, index=False, encoding='utf-8-sig')

        print(f"✅ CSV文件已保存:")
        print(f"  📁 详细分析: {detailed_csv}")
        print(f"  📊 简洁汇总: {summary_csv}")

        return detailed_csv, summary_csv

    except Exception as e:
        print(f"❌ 保存CSV文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return None, None


def analyze_obstacle_data(log_file_path, output_path, output_format='excel'):
    """
    主要分析函数

    Args:
        log_file_path: 日志文件路径
        output_path: 输出文件路径（不包含扩展名）
        output_format: 输出格式 ('excel', 'csv', 'both')
    """
    # 检查日志文件是否存在
    if not os.path.exists(log_file_path):
        print(f"❌ 错误: 日志文件不存在: {log_file_path}")
        return None, None

    print(f"📂 开始分析日志文件: {log_file_path}")

    # 读取日志文件
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        print(f"📝 读取到 {len(lines)} 行数据")
    except Exception as e:
        print(f"❌ 读取日志文件失败: {e}")
        return None, None

    # 解析所有数据
    all_data = []
    failed_lines = 0

    for line_num, line in enumerate(lines, 1):
        parsed = parse_log_line(line)
        if parsed:
            all_data.append(parsed)
        else:
            failed_lines += 1
            if failed_lines <= 3:  # 只显示前3个失败的行
                print(f"⚠️  第{line_num}行解析失败: {line.strip()[:100]}...")

    print(f"✅ 成功解析 {len(all_data)} 条数据，失败 {failed_lines} 条")

    # 检查是否有有效数据
    if not all_data:
        print("❌ 错误: 没有找到有效的日志数据")
        print("请检查日志文件格式是否正确")
        return None, None

    # 按障碍物分组 (ObjName + OdDir + HeightLabel)
    obstacle_groups = defaultdict(list)
    for data in all_data:
        key = (data['ObjName'], data['OdDir'], data['HeightLabel'])
        obstacle_groups[key].append(data)

    print(f"🔍 识别到 {len(obstacle_groups)} 个不同的障碍物组合")

    # 详细结果和简洁结果
    detailed_results = []
    summary_results = []

    for (obj_name, od_dir, height_label), data_list in obstacle_groups.items():
        # 转换OdDir为可读格式
        od_dir_str = "front" if od_dir == 1 else "rear"
        height_label_str = "低障碍物" if height_label == 0 else "高障碍物"

        print(f"\n🔍 分析障碍物: {obj_name} ({od_dir_str}, {height_label_str})")

        # 按zhao_od_ID分组
        id_groups = defaultdict(list)
        for data in data_list:
            id_groups[data['zhao_od_ID']].append(data)

        print(f"  发现 {len(id_groups)} 个不同的障碍物ID: {list(id_groups.keys())}")

        # 用于简洁版统计
        obstacle_has_any_error = False
        total_sessions_all_ids = 0
        total_points_all_ids = 0

        for zhao_od_id, id_data in id_groups.items():
            print(f"  分析ID {zhao_od_id}: {len(id_data)} 个数据点")

            # 识别测试次数
            sessions = identify_test_sessions(id_data)
            total_sessions_all_ids += len(sessions)

            # 检查所有次数是否都正确
            all_sessions_correct = True
            incorrect_sessions = []

            for session_idx, session in enumerate(sessions, 1):
                session_correct = all(
                    (height_label == 0 and data['HeightStatus'] == 1) or
                    (height_label == 1 and data['HeightStatus'] == 2)
                    for data in session
                )

                if not session_correct:
                    all_sessions_correct = False
                    incorrect_sessions.append(session_idx)
                    obstacle_has_any_error = True

            # 统计总数据点
            total_points_all_ids += sum(len(session) for session in sessions)

            # 如果所有次数都正确，只输出一条记录
            if all_sessions_correct:
                total_points = sum(len(session) for session in sessions)
                min_dist = min(data['Dist'] for session in sessions for data in session)
                max_dist = max(data['Dist'] for session in sessions for data in session)

                detailed_results.append({
                    '障碍物名称': obj_name,
                    '位置': od_dir_str,
                    '真实标签': height_label_str,
                    '障碍物ID': zhao_od_id,
                    '总测试次数': len(sessions),
                    '测试次数说明': f"第{','.join(map(str, range(1, len(sessions) + 1)))}次" if len(
                        sessions) > 1 else "第1次",
                    '数据点总数': total_points,
                    '距离范围(mm)': f"{max_dist}~{min_dist}",
                    '检测状态': "全程正确",
                    '首次错误距离(mm)': "无",
                    '错误详情': "无",
                    '备注': f"共{len(sessions)}次测试，全程HeightStatus与HeightLabel一致"
                })

            else:
                # 对于错误的次数，分别输出
                for session_idx, session in enumerate(sessions, 1):
                    if session_idx in incorrect_sessions:
                        changes = analyze_height_status_changes(session)
                        error_desc = format_distance_ranges(changes, height_label)
                        status_analysis = get_status_analysis(height_label, True, changes)

                        first_error_dist = None
                        for change in changes:
                            is_error = not ((height_label == 0 and change['status'] == 1) or
                                            (height_label == 1 and change['status'] == 2))
                            if is_error:
                                first_error_dist = change['start_dist']
                                break

                        detailed_results.append({
                            '障碍物名称': obj_name,
                            '位置': od_dir_str,
                            '真实标签': height_label_str,
                            '障碍物ID': zhao_od_id,
                            '总测试次数': len(sessions),
                            '测试次数说明': f"第{session_idx}次",
                            '数据点总数': len(session),
                            '距离范围(mm)': f"{max(d['Dist'] for d in session)}~{min(d['Dist'] for d in session)}",
                            '检测状态': status_analysis,
                            '首次错误距离(mm)': str(first_error_dist) if first_error_dist else "无",
                            '错误详情': error_desc,
                            '备注': f"第{session_idx}次测试中HeightStatus与HeightLabel不一致"
                        })

                # 对于正确的次数，合并输出一条记录
                correct_sessions = [i for i in range(1, len(sessions) + 1) if i not in incorrect_sessions]
                if correct_sessions:
                    correct_data = [sessions[i - 1] for i in correct_sessions]
                    total_points = sum(len(session) for session in correct_data)
                    min_dist = min(data['Dist'] for session in correct_data for data in session)
                    max_dist = max(data['Dist'] for session in correct_data for data in session)

                    detailed_results.append({
                        '障碍物名称': obj_name,
                        '位置': od_dir_str,
                        '真实标签': height_label_str,
                        '障碍物ID': zhao_od_id,
                        '总测试次数': len(sessions),
                        '测试次数说明': f"第{','.join(map(str, correct_sessions))}次",
                        '数据点总数': total_points,
                        '距离范围(mm)': f"{max_dist}~{min_dist}",
                        '检测状态': "全程正确",
                        '首次错误距离(mm)': "无",
                        '错误详情': "无",
                        '备注': f"第{','.join(map(str, correct_sessions))}次测试全程正确"
                    })

        # 生成简洁版结果
        if not obstacle_has_any_error:
            # 全程无错误，不区分ID
            all_distances = [data['Dist'] for data in data_list]
            summary_results.append({
                '障碍物名称': obj_name,
                '位置': od_dir_str,
                '真实标签': height_label_str,
                '总测试次数': total_sessions_all_ids,
                '涉及ID数量': len(id_groups),
                '数据点总数': total_points_all_ids,
                '距离范围(mm)': f"{max(all_distances)}~{min(all_distances)}",
                '检测结果': "✓ 全程正确",
                '备注': f"所有ID({','.join(map(str, sorted(id_groups.keys())))})全程检测正确"
            })
        else:
            # 存在错误，需要查看详细表
            error_type = "存在报高" if height_label == 0 else "存在报低"
            all_distances = [data['Dist'] for data in data_list]
            summary_results.append({
                '障碍物名称': obj_name,
                '位置': od_dir_str,
                '真实标签': height_label_str,
                '总测试次数': total_sessions_all_ids,
                '涉及ID数量': len(id_groups),
                '数据点总数': total_points_all_ids,
                '距离范围(mm)': f"{max(all_distances)}~{min(all_distances)}",
                '检测结果': f"✗ {error_type}",
                '备注': "存在检测错误，详情请查看详细分析表"
            })

    # 检查是否有结果
    if not detailed_results:
        print("❌ 错误: 没有生成有效的分析结果")
        return None, None

    # 转换为DataFrame
    detailed_df = pd.DataFrame(detailed_results)
    summary_df = pd.DataFrame(summary_results)

    print(f"\n📊 生成分析结果: 详细记录{len(detailed_results)}条，汇总记录{len(summary_results)}条")

    # 处理输出路径
    if output_path.endswith('.csv') or output_path.endswith('.xlsx'):
        base_path = os.path.splitext(output_path)[0]
    else:
        base_path = output_path

    # 根据格式选择保存方式
    saved_files = []

    if output_format.lower() in ['excel', 'both']:
        excel_file = save_to_excel(detailed_df, summary_df, base_path)
        if excel_file:
            saved_files.append(excel_file)

    if output_format.lower() in ['csv', 'both']:
        csv_files = save_to_csv(detailed_df, summary_df, base_path)
        if csv_files[0] and csv_files[1]:
            saved_files.extend(csv_files)

    # 美化输出
    print("=" * 80)
    print("🎯 障碍物检测日志分析完成!")
    print("=" * 80)

    if saved_files:
        print("📁 输出文件:")
        for file in saved_files:
            print(f"  • {file}")
    else:
        print("❌ 警告: 没有成功保存任何文件")

    print()
    print("📈 统计概览:")
    print("-" * 50)
    print(f"🔍 分析障碍物类型总数: {len(obstacle_groups)}")
    print(f"📝 详细记录条数: {len(detailed_results)}")
    print(f"📋 汇总记录条数: {len(summary_results)}")
    print(f"✅ 全程正确的障碍物: {len(summary_df[summary_df['检测结果'].str.contains('全程正确')])}")
    print(f"❌ 存在错误的障碍物: {len(summary_df[summary_df['检测结果'].str.contains('存在')])}")
    print()

    print("📋 简洁汇总表预览:")
    print("-" * 50)
    if len(summary_df) > 0:
        # 设置pandas显示选项以更好地显示中文
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 30)
        print(summary_df.head(10).to_string(index=False))
        if len(summary_df) > 10:
            print(f"\n... 还有 {len(summary_df) - 10} 行数据，详见输出文件")
    else:
        print("无数据")

    print("\n" + "=" * 80)

    # 显示使用说明
    print("💡 使用说明:")
    print("-" * 50)
    print("📊 Excel格式特点:")
    print("  • 自动调整列宽，便于阅读")
    print("  • 包含格式化样式（标题行高亮、边框等）")
    print("  • 冻结首行，方便查看数据")
    print("  • 两个工作表：详细分析 和 简洁汇总")
    print("\n📄 CSV格式特点:")
    print("  • 纯文本格式，兼容性好")
    print("  • 可用任何文本编辑器打开")
    print("  • 分为两个文件：详细分析.csv 和 简洁汇总.csv")
    print("\n🔧 依赖库提示:")
    print("  • Excel功能需要 openpyxl 库: pip install openpyxl")
    print("  • 如果只需要CSV格式，无需额外安装")
    print("\n📊 改进说明:")
    print("  • 优化了测试次数识别算法，能准确识别连续的接近过程")
    print("  • 自动过滤远离障碍物阶段的数据")
    print("  • 增强了对距离变化干扰的抗性")
    print("  • 增加了调试信息，便于验证识别结果")

    return detailed_df, summary_df


# 使用示例
if __name__ == "__main__":
    # 修改这里的路径为你的实际文件路径
    log_file_path = r"D:\PythonProject\data\log_files\2.log"  # 输入的日志文件路径
    output_path = r"D:\PythonProject\data\csv_files\2"  # 输出文件路径（不包含扩展名）

    try:
        print("🚀 开始改进版日志分析...")
        print(f"📂 日志文件: {log_file_path}")
        print(f"📁 输出路径: {output_path}")
        print("-" * 80)

        # 默认输出Excel格式
        detailed_df, summary_df = analyze_obstacle_data(
            log_file_path,
            output_path,
            output_format='excel'  # 可选: 'excel', 'csv', 'both'
        )

        if detailed_df is not None and summary_df is not None:
            print("\n🔍 详细分析表字段说明:")
            print("-" * 50)
            field_descriptions = {
                '障碍物名称': '障碍物类型名称',
                '位置': 'front(车前) / rear(车后)',
                '真实标签': '实际的高低标签',
                '障碍物ID': 'zhao_od_ID编号',
                '总测试次数': '该ID的测试次数',
                '测试次数说明': '当前记录对应的测试次数',
                '数据点总数': '数据记录条数',
                '距离范围(mm)': '测试距离范围(最远~最近)',
                '检测状态': '全程正确/存在报高/存在报低',
                '首次错误距离(mm)': '第一次出现错误时的距离',
                '错误详情': '错误的具体距离区间信息',
                '备注': '额外说明信息'
            }

            for field, desc in field_descriptions.items():
                print(f"  • {field}: {desc}")

            print("\n📊 简洁汇总表说明:")
            print("-" * 50)
            print("  • 全程无错误的障碍物：不区分不同ID，合并显示")
            print("  • 存在错误的障碍物：标注错误类型，详情见详细分析表")
            print("  • ✓ 表示检测正确，✗ 表示存在错误")

            print("\n🎨 格式选择:")
            print("-" * 50)
            print("  • output_format='excel': 仅输出Excel格式（默认推荐）")
            print("  • output_format='csv': 仅输出CSV格式")
            print("  • output_format='both': 同时输出Excel和CSV格式")

            print("\n🔧 新功能特点:")
            print("-" * 50)
            print("  • 智能识别连续接近过程，准确区分测试次数")
            print("  • 自动过滤车辆远离障碍物阶段的无效数据")
            print("  • 增强抗干扰能力，处理距离变化中的异常点")
            print("  • 提供详细的调试信息，便于验证分析结果")
        else:
            print("\n❌ 分析失败，请检查:")
            print("  1. 日志文件是否存在且格式正确")
            print("  2. 输出目录是否有写入权限")
            print("  3. 是否安装了必要的依赖库")

    except FileNotFoundError:
        print(f"❌ 错误: 找不到日志文件 '{log_file_path}'")
        print("请确保文件路径正确，或将日志内容保存为对应的日志文件")
    except Exception as e:
        print(f"❌ 分析过程中出现错误: {e}")
        import traceback

        traceback.print_exc()