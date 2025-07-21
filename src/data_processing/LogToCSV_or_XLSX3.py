import re
import pandas as pd
from collections import defaultdict, OrderedDict
import numpy as np
import os
from pathlib import Path


def parse_log_line(line):
    """解析单行日志数据"""
    pattern = r'zhao_od_ID:(\d+), ObjName:([^,]+), HeightLabel:(\d+), OdDir:(\d+), P1\(([^)]+)\), P2\(([^)]+)\), HeightProb:(\d+), HeightStatus:(\d+), Dist:(\d+), PosDeDis1:(\d+), l_TrainResult:(\d+), probability:(\d+)'
    match = re.match(pattern, line.strip())

    if match:
        return {
            'zhao_od_ID': int(match.group(1)),
            'ObjName': match.group(2),
            'HeightLabel': int(match.group(3)),
            'OdDir': int(match.group(4)),
            'P1': match.group(5),
            'P2': match.group(6),
            'HeightProb': int(match.group(7)),
            'HeightStatus': int(match.group(8)),
            'Dist': int(match.group(9)),
            'PosDeDis1': int(match.group(10)),
            'l_TrainResult': int(match.group(11)),
            'probability': int(match.group(12))
        }
    return None


def identify_sessions_by_distance_only(all_data):
    """
    完全重写的采样识别算法：
    基于距离变化的全局分析，识别完整的接近和远离过程
    """
    if len(all_data) <= 1:
        return [all_data], [(data, 1, 'approach') for data in all_data]

    distances = [d['Dist'] for d in all_data]
    labeled_data = []
    approach_sessions = []

    print(f"    调试：开始分析 {len(distances)} 个数据点，距离范围 {min(distances)}-{max(distances)}")

    # 第一步：使用滑动窗口计算每个点的局部趋势
    def calculate_local_trends(distances, window_size=8):
        """计算每个点的局部趋势值"""
        trends = []

        for i in range(len(distances)):
            # 向前和向后各取window_size//2个点
            half_window = window_size // 2
            start_idx = max(0, i - half_window)
            end_idx = min(len(distances), i + half_window + 1)

            local_distances = distances[start_idx:end_idx]
            if len(local_distances) < 3:
                trends.append(0)  # 数据点太少，趋势为0
                continue

            # 计算线性回归斜率作为趋势值
            x = list(range(len(local_distances)))
            y = local_distances
            n = len(x)

            sum_x = sum(x)
            sum_y = sum(y)
            sum_xy = sum(x[i] * y[i] for i in range(n))
            sum_x2 = sum(x[i] * x[i] for i in range(n))

            # 斜率 = (n*Σxy - Σx*Σy) / (n*Σx² - (Σx)²)
            denominator = n * sum_x2 - sum_x * sum_x
            if denominator == 0:
                slope = 0
            else:
                slope = (n * sum_xy - sum_x * sum_y) / denominator

            trends.append(slope)

        return trends

    # 计算局部趋势
    trends = calculate_local_trends(distances)

    # 第二步：基于趋势识别接近和远离的阶段
    def identify_phases(trends, distances, smoothing_window=5):
        """识别接近和远离阶段"""
        # 对趋势进行平滑处理
        smoothed_trends = []
        for i in range(len(trends)):
            start = max(0, i - smoothing_window // 2)
            end = min(len(trends), i + smoothing_window // 2 + 1)
            smoothed_trends.append(sum(trends[start:end]) / (end - start))

        phases = []
        current_phase = 'approach'  # 假设开始是接近阶段
        phase_start = 0

        # 设置阈值
        TREND_THRESHOLD = 5.0  # 趋势阈值，正值表示上升，负值表示下降
        MIN_PHASE_LENGTH = 8  # 最小阶段长度

        for i in range(1, len(smoothed_trends)):
            should_switch = False

            if current_phase == 'approach':
                # 接近阶段：寻找明显的上升趋势（远离）
                if smoothed_trends[i] > TREND_THRESHOLD and i - phase_start >= MIN_PHASE_LENGTH:
                    # 检查后续几个点是否确实在上升
                    look_ahead = min(len(distances) - i, 5)
                    if look_ahead >= 3:
                        future_trend = sum(smoothed_trends[i:i + look_ahead]) / look_ahead
                        if future_trend > TREND_THRESHOLD * 0.5:
                            should_switch = True
            else:  # retreat
                # 远离阶段：寻找明显的下降趋势（接近）
                if smoothed_trends[i] < -TREND_THRESHOLD and i - phase_start >= MIN_PHASE_LENGTH:
                    # 检查后续几个点是否确实在下降
                    look_ahead = min(len(distances) - i, 5)
                    if look_ahead >= 3:
                        future_trend = sum(smoothed_trends[i:i + look_ahead]) / look_ahead
                        if future_trend < -TREND_THRESHOLD * 0.5:
                            should_switch = True

            if should_switch:
                # 记录当前阶段
                phases.append({
                    'type': current_phase,
                    'start': phase_start,
                    'end': i - 1,
                    'length': i - phase_start
                })

                # 切换到新阶段
                current_phase = 'retreat' if current_phase == 'approach' else 'approach'
                phase_start = i

        # 添加最后一个阶段
        phases.append({
            'type': current_phase,
            'start': phase_start,
            'end': len(distances) - 1,
            'length': len(distances) - phase_start
        })

        return phases

    # 识别阶段
    phases = identify_phases(trends, distances)

    print(f"    调试：识别到 {len(phases)} 个阶段")
    for i, phase in enumerate(phases):
        phase_type = "接近" if phase['type'] == 'approach' else "远离"
        start_dist = distances[phase['start']]
        end_dist = distances[phase['end']]
        print(f"      第{i + 1}个阶段：{phase_type}，长度={phase['length']}，距离={start_dist}->{end_dist}")

    # 第三步：生成标记数据和提取接近会话
    approach_count = 0
    retreat_count = 0

    for phase in phases:
        if phase['type'] == 'approach':
            approach_count += 1
            session_data = all_data[phase['start']:phase['end'] + 1]

            # 只有足够长的接近阶段才被记录为有效会话
            if len(session_data) >= 5:
                approach_sessions.append(session_data)

            # 标记数据
            for i in range(phase['start'], phase['end'] + 1):
                labeled_data.append((all_data[i], approach_count, 'approach'))
        else:  # retreat
            retreat_count += 1
            # 标记远离数据
            for i in range(phase['start'], phase['end'] + 1):
                labeled_data.append((all_data[i], retreat_count, 'retreat'))

    # 为每个接近会话按距离排序（从大到小）
    for session in approach_sessions:
        session.sort(key=lambda x: x['Dist'], reverse=True)

    print(f"    调试：提取到 {len(approach_sessions)} 个有效接近会话")

    return approach_sessions, labeled_data


def analyze_height_status_changes(session_data):
    """分析单次测试中HeightStatus的变化"""
    session_data.sort(key=lambda x: x['Dist'], reverse=True)

    changes = []
    if not session_data:
        return changes

    current_status = session_data[0]['HeightStatus']
    change_start_idx = 0

    for i, data in enumerate(session_data):
        if data['HeightStatus'] != current_status:
            changes.append({
                'start_idx': change_start_idx,
                'end_idx': i - 1,
                'status': current_status,
                'start_dist': session_data[change_start_idx]['Dist'],
                'end_dist': session_data[i - 1]['Dist']
            })
            current_status = data['HeightStatus']
            change_start_idx = i

    if change_start_idx < len(session_data):
        changes.append({
            'start_idx': change_start_idx,
            'end_idx': len(session_data) - 1,
            'status': current_status,
            'start_dist': session_data[change_start_idx]['Dist'],
            'end_dist': session_data[-1]['Dist']
        })

    return changes


def format_distance_ranges(changes, height_label):
    """格式化距离区间信息"""
    if not changes:
        return "无数据"

    wrong_ranges = []
    correct_ranges = []
    first_wrong_dist = None

    for change in changes:
        status_desc = "报高" if change['status'] == 2 else "报低"
        range_desc = f"Dist:{change['start_dist']}-{change['end_dist']}"

        is_correct = (height_label == 0 and change['status'] == 1) or (height_label == 1 and change['status'] == 2)

        if is_correct:
            correct_ranges.append(f"{range_desc}({status_desc})")
        else:
            wrong_ranges.append(f"{range_desc}({status_desc})")
            if first_wrong_dist is None:
                first_wrong_dist = change['start_dist']

    descriptions = []
    if first_wrong_dist is not None:
        descriptions.append(f"首次错误报告距离:{first_wrong_dist}")

    if wrong_ranges:
        descriptions.append(f"错误区间:{'; '.join(wrong_ranges)}")

    if correct_ranges:
        descriptions.append(f"正确区间:{'; '.join(correct_ranges)}")

    if len(changes) > 1:
        last_change = changes[-1]
        last_is_correct = (height_label == 0 and last_change['status'] == 1) or (
                height_label == 1 and last_change['status'] == 2)
        if last_is_correct and last_change['end_dist'] < last_change['start_dist'] * 0.7:
            descriptions.append(f"Dist<{last_change['start_dist']}后全程{'报低' if height_label == 0 else '报高'}")

    return "; ".join(descriptions) if descriptions else "全程正确"


def get_status_analysis(height_label, has_error, changes):
    """获取状态分析描述"""
    if not has_error:
        return "全程正确"

    for change in changes:
        is_error = not ((height_label == 0 and change['status'] == 1) or
                        (height_label == 1 and change['status'] == 2))
        if is_error:
            if height_label == 0 and change['status'] == 2:
                return "存在报高"
            elif height_label == 1 and change['status'] == 1:
                return "存在报低"

    return "存在错误"


def create_raw_data_sheet(labeled_data):
    """创建原始数据表，显示每条数据的分组情况"""
    raw_results = []

    for data, count, phase in labeled_data:
        od_dir_str = "front" if data['OdDir'] == 1 else "rear"
        height_label_str = "低障碍物" if data['HeightLabel'] == 0 else "高障碍物"
        height_status_str = "报低" if data['HeightStatus'] == 1 else "报高"
        phase_str = f"第{count}次接近" if phase == 'approach' else f"第{count}次远离"

        raw_results.append({
            '障碍物名称': data['ObjName'],
            '位置': od_dir_str,
            '真实标签': height_label_str,
            '障碍物ID': data['zhao_od_ID'],
            'P1坐标': data['P1'],
            'P2坐标': data['P2'],
            '距离Dist': data['Dist'],
            '雷达距离': data['PosDeDis1'],
            '检测状态': height_status_str,
            '置信度': data['HeightProb'],
            '模型预测': data['l_TrainResult'],
            '模型概率': data['probability'],
            '检测阶段': phase_str,
            '是否正确': "✓" if ((data['HeightLabel'] == 0 and data['HeightStatus'] == 1) or
                                (data['HeightLabel'] == 1 and data['HeightStatus'] == 2)) else "✗"
        })

    return pd.DataFrame(raw_results)


def ensure_directory_exists(file_path):
    """确保文件所在目录存在"""
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        try:
            os.makedirs(directory, exist_ok=True)
            print(f"✅ 创建目录: {directory}")
        except Exception as e:
            print(f"❌ 创建目录失败: {e}")
            return False
    return True


def save_to_excel(detailed_df, summary_df, raw_df, base_path, include_raw=False):
    """保存为Excel格式，自动调整列宽"""
    excel_path = f"{base_path}_分析结果.xlsx"

    if not ensure_directory_exists(excel_path):
        return None

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            detailed_df.to_excel(writer, sheet_name='详细分析', index=False)
            summary_df.to_excel(writer, sheet_name='简洁汇总', index=False)

            if include_raw and raw_df is not None:
                raw_df.to_excel(writer, sheet_name='原始数据分组', index=False)

            workbook = writer.book
            detailed_sheet = workbook['详细分析']
            adjust_column_width(detailed_sheet, detailed_df)

            summary_sheet = workbook['简洁汇总']
            adjust_column_width(summary_sheet, summary_df)

            if include_raw and raw_df is not None:
                raw_sheet = workbook['原始数据分组']
                adjust_column_width(raw_sheet, raw_df)

            format_excel_sheets(workbook)

        sheet_info = "包含原始数据分组表" if include_raw else "不包含原始数据表"
        print(f"✅ Excel文件已保存: {excel_path} ({sheet_info})")
        return excel_path

    except ImportError:
        print("⚠️  警告: 未安装 openpyxl 库，无法生成Excel文件")
        return None
    except Exception as e:
        print(f"❌ 保存Excel文件时出错: {e}")
        return None


def adjust_column_width(sheet, df):
    """自动调整列宽"""
    try:
        from openpyxl.utils import get_column_letter

        for col_idx, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(column))

            for value in df[column]:
                if pd.notna(value):
                    str_value = str(value)
                    length = len(str_value) + sum(1 for char in str_value if ord(char) > 127)
                    max_length = max(max_length, length)

            adjusted_width = min(max(max_length + 2, 10), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    except ImportError:
        print("⚠️  警告: 无法调整列宽，请确保安装了 openpyxl 库")
    except Exception as e:
        print(f"⚠️  调整列宽时出现问题: {e}")


def format_excel_sheets(workbook):
    """格式化Excel工作表"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            sheet.freeze_panes = 'A2'

    except ImportError:
        print("⚠️  警告: 无法应用Excel格式化")
    except Exception as e:
        print(f"⚠️  格式化Excel时出现问题: {e}")


def save_to_csv(detailed_df, summary_df, raw_df, base_path, include_raw=False):
    """保存为CSV格式"""
    detailed_csv = f"{base_path}_详细分析.csv"
    summary_csv = f"{base_path}_简洁汇总.csv"
    raw_csv = f"{base_path}_原始数据分组.csv"

    if not ensure_directory_exists(detailed_csv) or not ensure_directory_exists(summary_csv):
        return None, None, None

    try:
        detailed_df.to_csv(detailed_csv, index=False, encoding='utf-8-sig')
        summary_df.to_csv(summary_csv, index=False, encoding='utf-8-sig')

        saved_files = [detailed_csv, summary_csv]

        if include_raw and raw_df is not None:
            raw_df.to_csv(raw_csv, index=False, encoding='utf-8-sig')
            saved_files.append(raw_csv)

        print(f"✅ CSV文件已保存:")
        print(f"  📁 详细分析: {detailed_csv}")
        print(f"  📊 简洁汇总: {summary_csv}")
        if include_raw and raw_df is not None:
            print(f"  📋 原始数据分组: {raw_csv}")

        return tuple(saved_files)

    except Exception as e:
        print(f"❌ 保存CSV文件时出错: {e}")
        return None, None, None


def analyze_obstacle_data(log_file_path, output_path, output_format='excel', include_raw_data=False):
    """
    主要分析函数
    """
    if not os.path.exists(log_file_path):
        print(f"❌ 错误: 日志文件不存在: {log_file_path}")
        return None, None, None

    print(f"📂 开始分析日志文件: {log_file_path}")
    print(f"🔧 原始数据分组表: {'启用' if include_raw_data else '禁用'}")

    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        print(f"📝 读取到 {len(lines)} 行数据")
    except Exception as e:
        print(f"❌ 读取日志文件失败: {e}")
        return None, None, None

    all_data = []
    failed_lines = 0

    for line_num, line in enumerate(lines, 1):
        parsed = parse_log_line(line)
        if parsed:
            all_data.append(parsed)
        else:
            failed_lines += 1
            if failed_lines <= 3:
                print(f"⚠️  第{line_num}行解析失败: {line.strip()[:100]}...")

    print(f"✅ 成功解析 {len(all_data)} 条数据，失败 {failed_lines} 条")

    if not all_data:
        print("❌ 错误: 没有找到有效的日志数据")
        return None, None, None

    obstacle_groups = defaultdict(list)
    for data in all_data:
        key = (data['ObjName'], data['OdDir'], data['HeightLabel'])
        obstacle_groups[key].append(data)

    print(f"🔍 识别到 {len(obstacle_groups)} 个不同的障碍物组合")

    detailed_results = []
    summary_results = []
    all_labeled_data = []

    for (obj_name, od_dir, height_label), data_list in obstacle_groups.items():
        od_dir_str = "front" if od_dir == 1 else "rear"
        height_label_str = "低障碍物" if height_label == 0 else "高障碍物"

        print(f"\n🔍 分析障碍物: {obj_name} ({od_dir_str}, {height_label_str})")

        approach_sessions, labeled_data = identify_sessions_by_distance_only(data_list)
        all_labeled_data.extend(labeled_data)

        print(f"  识别出 {len(approach_sessions)} 次接近过程")

        id_groups = defaultdict(list)
        for session_idx, session in enumerate(approach_sessions, 1):
            for data in session:
                id_groups[data['zhao_od_ID']].append((data, session_idx))

        print(f"  涉及 {len(id_groups)} 个不同的障碍物ID: {list(id_groups.keys())}")

        obstacle_has_any_error = False
        total_sessions_all_ids = len(approach_sessions)
        total_points_all_ids = sum(len(session) for session in approach_sessions)

        for zhao_od_id, id_data_with_session in id_groups.items():
            print(f"  分析ID {zhao_od_id}: {len(id_data_with_session)} 个数据点")

            session_groups = defaultdict(list)
            for data, session_idx in id_data_with_session:
                session_groups[session_idx].append(data)

            all_sessions_correct = True
            incorrect_sessions = []

            for session_idx, session_data in session_groups.items():
                session_correct = all(
                    (height_label == 0 and data['HeightStatus'] == 1) or
                    (height_label == 1 and data['HeightStatus'] == 2)
                    for data in session_data
                )

                if not session_correct:
                    all_sessions_correct = False
                    incorrect_sessions.append(session_idx)
                    obstacle_has_any_error = True

            if all_sessions_correct:
                total_points = sum(len(session_data) for session_data in session_groups.values())
                all_distances = [data['Dist'] for session_data in session_groups.values() for data in session_data]
                min_dist = min(all_distances)
                max_dist = max(all_distances)

                detailed_results.append({
                    '障碍物名称': obj_name,
                    '位置': od_dir_str,
                    '真实标签': height_label_str,
                    '障碍物ID': zhao_od_id,
                    '总测试次数': len(session_groups),
                    '测试次数说明': f"第{','.join(map(str, sorted(session_groups.keys())))}次" if len(
                        session_groups) > 1 else "第1次",
                    '数据点总数': total_points,
                    '距离范围(mm)': f"{max_dist}~{min_dist}",
                    '检测状态': "全程正确",
                    '首次错误距离(mm)': "无",
                    '错误详情': "无",
                    '备注': f"共{len(session_groups)}次测试，全程HeightStatus与HeightLabel一致"
                })

            else:
                # 对于错误的次数，分别输出
                for session_idx in incorrect_sessions:
                    session_data = session_groups[session_idx]
                    changes = analyze_height_status_changes(session_data)
                    error_desc = format_distance_ranges(changes, height_label)
                    status_analysis = get_status_analysis(height_label, True, changes)

                    first_error_dist = None
                    for change in changes:
                        is_error = not ((height_label == 0 and change['status'] == 1) or
                                        (height_label == 1 and change['status'] == 2))
                        if is_error:
                            first_error_dist = change['start_dist']
                            break

                    detailed_results.append({
                        '障碍物名称': obj_name,
                        '位置': od_dir_str,
                        '真实标签': height_label_str,
                        '障碍物ID': zhao_od_id,
                        '总测试次数': len(session_groups),
                        '测试次数说明': f"第{session_idx}次",
                        '数据点总数': len(session_data),
                        '距离范围(mm)': f"{max(d['Dist'] for d in session_data)}~{min(d['Dist'] for d in session_data)}",
                        '检测状态': status_analysis,
                        '首次错误距离(mm)': str(first_error_dist) if first_error_dist else "无",
                        '错误详情': error_desc,
                        '备注': f"第{session_idx}次测试中HeightStatus与HeightLabel不一致"
                    })

                # 对于正确的次数，合并输出一条记录
                correct_sessions = [i for i in session_groups.keys() if i not in incorrect_sessions]
                if correct_sessions:
                    correct_data = [session_groups[i] for i in correct_sessions]
                    total_points = sum(len(session_data) for session_data in correct_data)
                    all_distances = [data['Dist'] for session_data in correct_data for data in session_data]
                    min_dist = min(all_distances)
                    max_dist = max(all_distances)

                    detailed_results.append({
                        '障碍物名称': obj_name,
                        '位置': od_dir_str,
                        '真实标签': height_label_str,
                        '障碍物ID': zhao_od_id,
                        '总测试次数': len(session_groups),
                        '测试次数说明': f"第{','.join(map(str, correct_sessions))}次",
                        '数据点总数': total_points,
                        '距离范围(mm)': f"{max_dist}~{min_dist}",
                        '检测状态': "全程正确",
                        '首次错误距离(mm)': "无",
                        '错误详情': "无",
                        '备注': f"第{','.join(map(str, correct_sessions))}次测试全程正确"
                    })

        # 生成简洁版结果
        if not obstacle_has_any_error:
            all_distances = [data['Dist'] for data in data_list]
            summary_results.append({
                '障碍物名称': obj_name,
                '位置': od_dir_str,
                '真实标签': height_label_str,
                '总测试次数': total_sessions_all_ids,
                '涉及ID数量': len(id_groups),
                '数据点总数': total_points_all_ids,
                '距离范围(mm)': f"{max(all_distances)}~{min(all_distances)}",
                '检测结果': "✓ 全程正确",
                '备注': f"所有ID({','.join(map(str, sorted(id_groups.keys())))})全程检测正确"
            })
        else:
            error_type = "存在报高" if height_label == 0 else "存在报低"
            all_distances = [data['Dist'] for data in data_list]
            summary_results.append({
                '障碍物名称': obj_name,
                '位置': od_dir_str,
                '真实标签': height_label_str,
                '总测试次数': total_sessions_all_ids,
                '涉及ID数量': len(id_groups),
                '数据点总数': total_points_all_ids,
                '距离范围(mm)': f"{max(all_distances)}~{min(all_distances)}",
                '检测结果': f"✗ {error_type}",
                '备注': "存在检测错误，详情请查看详细分析表"
            })

    if not detailed_results:
        print("❌ 错误: 没有生成有效的分析结果")
        return None, None, None

    detailed_df = pd.DataFrame(detailed_results)
    summary_df = pd.DataFrame(summary_results)

    raw_df = None
    if include_raw_data:
        raw_df = create_raw_data_sheet(all_labeled_data)
        print(f"📋 生成原始数据分组表: {len(raw_df)}条记录")

    print(f"\n📊 生成分析结果: 详细记录{len(detailed_results)}条，汇总记录{len(summary_results)}条")

    if output_path.endswith('.csv') or output_path.endswith('.xlsx'):
        base_path = os.path.splitext(output_path)[0]
    else:
        base_path = output_path

    saved_files = []

    if output_format.lower() in ['excel', 'both']:
        excel_file = save_to_excel(detailed_df, summary_df, raw_df, base_path, include_raw_data)
        if excel_file:
            saved_files.append(excel_file)

    if output_format.lower() in ['csv', 'both']:
        csv_files = save_to_csv(detailed_df, summary_df, raw_df, base_path, include_raw_data)
        if csv_files and csv_files[0]:
            saved_files.extend([f for f in csv_files if f])

    print("=" * 80)
    print("🎯 障碍物检测日志分析完成!")
    print("=" * 80)

    if saved_files:
        print("📁 输出文件:")
        for file in saved_files:
            print(f"  • {file}")
    else:
        print("❌ 警告: 没有成功保存任何文件")

    print()
    print("📈 统计概览:")
    print("-" * 50)
    print(f"🔍 分析障碍物类型总数: {len(obstacle_groups)}")
    print(f"📝 详细记录条数: {len(detailed_results)}")
    print(f"📋 汇总记录条数: {len(summary_results)}")
    print(f"✅ 全程正确的障碍物: {len(summary_df[summary_df['检测结果'].str.contains('全程正确')])}")
    print(f"❌ 存在错误的障碍物: {len(summary_df[summary_df['检测结果'].str.contains('存在')])}")
    if include_raw_data and raw_df is not None:
        print(f"📋 原始数据记录条数: {len(raw_df)}")

    print("\n📋 简洁汇总表预览:")
    print("-" * 50)
    if len(summary_df) > 0:
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 30)
        print(summary_df.head(10).to_string(index=False))
        if len(summary_df) > 10:
            print(f"\n... 还有 {len(summary_df) - 10} 行数据，详见输出文件")
    else:
        print("无数据")

    print("\n" + "=" * 80)

    print("💡 全新算法说明:")
    print("-" * 50)
    print("🔧 核心改进:")
    print("  • 基于线性回归的局部趋势计算")
    print("  • 全局阶段识别而非逐点判断")
    print("  • 使用平滑处理减少噪声影响")
    print("  • 明确的阈值和最小阶段长度限制")
    print("  • 前瞻性验证避免误切换")
    print("\n📊 算法步骤:")
    print("  1. 计算每个点的局部趋势值（线性回归斜率）")
    print("  2. 对趋势进行平滑处理")
    print("  3. 基于趋势阈值识别接近/远离阶段")
    print("  4. 验证阶段切换的合理性")
    print("  5. 提取有效的接近会话")

    return detailed_df, summary_df, raw_df


# 使用示例
if __name__ == "__main__":
    log_file_path = r"D:\PythonProject\data\log_files\2.log"
    output_path = r"D:\PythonProject\data\csv_files\2"

    try:
        print("🚀 开始全新算法日志分析...")
        print(f"📂 日志文件: {log_file_path}")
        print(f"📁 输出路径: {output_path}")
        print("-" * 80)

        include_raw_data = True

        detailed_df, summary_df, raw_df = analyze_obstacle_data(
            log_file_path,
            output_path,
            output_format='excel',
            include_raw_data=include_raw_data
        )

        if detailed_df is not None and summary_df is not None:
            print("\n🔍 详细分析表字段说明:")
            print("-" * 50)
            field_descriptions = {
                '障碍物名称': '障碍物类型名称',
                '位置': 'front(车前) / rear(车后)',
                '真实标签': '实际的高低标签',
                '障碍物ID': 'zhao_od_ID编号',
                '总测试次数': '该ID在所有接近过程中的测试次数',
                '测试次数说明': '当前记录对应的测试次数',
                '数据点总数': '数据记录条数',
                '距离范围(mm)': '测试距离范围(最远~最近)',
                '检测状态': '全程正确/存在报高/存在报低',
                '首次错误距离(mm)': '第一次出现错误时的距离',
                '错误详情': '错误的具体距离区间信息',
                '备注': '额外说明信息'
            }

            for field, desc in field_descriptions.items():
                print(f"  • {field}: {desc}")

            if include_raw_data and raw_df is not None:
                print("\n📋 原始数据分组表字段说明:")
                print("-" * 50)
                raw_field_descriptions = {
                    '障碍物名称': '障碍物类型名称',
                    '位置': 'front(车前) / rear(车后)',
                    '真实标签': '实际的高低标签',
                    '障碍物ID': 'zhao_od_ID编号',
                    'P1坐标': '障碍物坐标点1',
                    'P2坐标': '障碍物坐标点2',
                    '距离Dist': '车辆到障碍物的距离',
                    '雷达距离': '雷达直接回波距离',
                    '检测状态': '系统报告的高低状态',
                    '置信度': '系统置信度',
                    '模型预测': '模型预测结果',
                    '模型概率': '模型预测概率值',
                    '检测阶段': '第几次接近或远离过程',
                    '是否正确': '检测结果是否与真实标签一致'
                }

                for field, desc in raw_field_descriptions.items():
                    print(f"  • {field}: {desc}")

            print("\n📊 核心算法突破:")
            print("-" * 50)
            print("  • 完全重写：从逐点判断改为全局阶段分析")
            print("  • 数学模型：使用线性回归计算趋势强度")
            print("  • 噪声处理：多层平滑减少干扰")
            print("  • 智能切换：前瞻性验证防止误判")
            print("  • 可靠性：明确的阈值和最小长度限制")

        else:
            print("\n❌ 分析失败，请检查:")
            print("  1. 日志文件是否存在且格式正确")
            print("  2. 输出目录是否有写入权限")
            print("  3. 是否安装了必要的依赖库")

    except FileNotFoundError:
        print(f"❌ 错误: 找不到日志文件 '{log_file_path}'")
        print("请确保文件路径正确，或将日志内容保存为对应的日志文件")
    except Exception as e:
        print(f"❌ 分析过程中出现错误: {e}")
        import traceback

        traceback.print_exc()